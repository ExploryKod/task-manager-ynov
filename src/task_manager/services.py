import smtplib
import csv
import re
import json
import xml.etree.ElementTree as ET
from datetime import datetime
from typing import List, Dict, Any, Optional
from .task import Task, Status, Priority

# Import conditionnel pour Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class EmailService:
    """Service d'envoi d'emails (à mocker dans les tests)"""

    MAX_EMAIL_LENGTH = 320
    
    def __init__(self, smtp_server: str = "smtp.gmail.com", port: int = 587) -> None:
        self.smtp_server: str = smtp_server
        self.port: int = port
        self.sent_emails: List[Dict[str, Any]] = []

    def send_task_reminder(self, email: str, task_title: str, due_date: datetime) -> bool:
        self._validate_email(email)
        self._validate_task_title(task_title)
        
        if not isinstance(due_date, datetime):
            raise TypeError(f"Due date must be a datetime object, got {type(due_date)}")
        
        email_data = {
            "to": email,
            "subject": f"Task Reminder: {task_title}",
            "body": f"Reminder: Your task '{task_title}' is due on {due_date.strftime('%Y-%m-%d %H:%M')}",
            "sent_at": datetime.now().isoformat(),
            "type": "reminder"
        }
        
        self._simulate_email_sending(email_data)
        return True

    def send_completion_notification(self, email: str, task_title: str) -> bool:
        self._validate_email(email)
        self._validate_task_title(task_title)
        
        email_data = {
            "to": email,
            "subject": f"Task Completed: {task_title}",
            "body": f"Congratulations! Your task '{task_title}' has been marked as completed.",
            "sent_at": datetime.now().isoformat(),
            "type": "completion"
        }
        
        self._simulate_email_sending(email_data)
        return True

    def _validate_email(self, email: str) -> None:
        if not isinstance(email, str):
            raise TypeError(f"Email must be a string, got {type(email)}")
        
        if not email or not email.strip():
            raise ValueError("Email cannot be empty")
        
        email = email.strip()
        
        if len(email) > self.MAX_EMAIL_LENGTH:
            raise ValueError(f"Email too long: {len(email)} characters. Maximum: {self.MAX_EMAIL_LENGTH}")
        
        if '@' not in email:
            raise ValueError(f"Invalid email format: missing '@' symbol in '{email}'")
        
        local_part, domain_part = email.split('@', 1)
        
        if not local_part:
            raise ValueError(f"Invalid email format: missing local part in '{email}'")
        
        if not domain_part:
            raise ValueError(f"Invalid email format: missing domain in '{email}'")
        
        if '.' not in domain_part:
            raise ValueError(f"Invalid email format: missing extension in '{email}'")
        
        if re.search(r'[<>"\s]', email):
            raise ValueError(f"Invalid email format: contains invalid characters in '{email}'")

    def _validate_task_title(self, task_title: str) -> None:
        if not isinstance(task_title, str):
            raise TypeError(f"Task title must be a string, got {type(task_title)}")
        
        if not task_title or not task_title.strip():
            raise ValueError("Task title cannot be empty")

    def _simulate_email_sending(self, email_data: Dict[str, Any]) -> None:
        self.sent_emails.append(email_data)
        print(f"[EMAIL SENT] {email_data['subject']} to {email_data['to']}")

    def get_sent_emails(self) -> List[Dict[str, Any]]:
        return self.sent_emails.copy()

    def clear_sent_emails(self) -> None:
        self.sent_emails.clear()


class ReportService:
    """Service de génération de rapports"""

    def generate_daily_report(
        self, 
        tasks: List[Task], 
        date: Optional[datetime] = None
    ) -> Dict[str, Any]:
        if not isinstance(tasks, list):
            raise TypeError(f"Tasks must be a list, got {type(tasks)}")
        
        if len(tasks) == 0:
            raise ValueError("Cannot generate report: no tasks found. At least one task is required.")
        
        report_date = date if date is not None else datetime.now()
        
        if not isinstance(report_date, datetime):
            raise TypeError(f"Date must be a datetime object, got {type(report_date)}")
        
        tasks_for_date = self._filter_tasks_by_date(tasks, report_date)
        
        completed_today = len([t for t in tasks_for_date if t.status == Status.DONE])
        created_today = len([t for t in tasks_for_date if t.created_at.date() == report_date.date()])
        
        priority_breakdown = {
            "urgent": len([t for t in tasks_for_date if t.priority == Priority.URGENT]),
            "high": len([t for t in tasks_for_date if t.priority == Priority.HIGH]),
            "medium": len([t for t in tasks_for_date if t.priority == Priority.MEDIUM]),
            "low": len([t for t in tasks_for_date if t.priority == Priority.LOW])
        }
        
        status_breakdown = {
            "todo": len([t for t in tasks_for_date if t.status == Status.TODO]),
            "in_progress": len([t for t in tasks_for_date if t.status == Status.IN_PROGRESS]),
            "done": len([t for t in tasks_for_date if t.status == Status.DONE]),
            "cancelled": len([t for t in tasks_for_date if t.status == Status.CANCELLED])
        }
        
        return {
            "report_date": report_date.isoformat(),
            "total_tasks": len(tasks),
            "tasks_for_date": len(tasks_for_date),
            "completed_today": completed_today,
            "created_today": created_today,
            "completion_rate_today": (completed_today / len(tasks_for_date) * 100) if tasks_for_date else 0,
            "priority_breakdown": priority_breakdown,
            "status_breakdown": status_breakdown,
            "generated_at": datetime.now().isoformat(),
            "summary": f"Daily report for {report_date.strftime('%Y-%m-%d')}: {completed_today} tasks completed, {created_today} tasks created"
        }

    def export_tasks_csv(self, tasks: List[Task], filename: str) -> bool:
        if not isinstance(tasks, list):
            raise TypeError(f"Tasks must be a list, got {type(tasks)}")
        
        if not isinstance(filename, str):
            raise TypeError(f"Filename must be a string, got {type(filename)}")
        
        if not filename or not filename.strip():
            raise ValueError("Filename cannot be empty")
        
        if len(tasks) == 0:
            raise ValueError("Cannot export CSV: no tasks to export. At least one task is required.")
        
        filename = filename.strip()
        
        if not filename.endswith('.csv'):
            filename += '.csv'
        
        try:
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = [
                    'id', 'title', 'description', 'priority', 'status',
                    'created_at', 'completed_at', 'project_id'
                ]
                
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                
                for task in tasks:
                    row = {
                        'id': task.id,
                        'title': task.title,
                        'description': task.description,
                        'priority': task.priority.value,
                        'status': task.status.value,
                        'created_at': task.created_at.isoformat(),
                        'completed_at': task.completed_at.isoformat() if task.completed_at else '',
                        'project_id': task.project_id if task.project_id else ''
                    }
                    writer.writerow(row)
            
            return True
            
        except PermissionError as e:
            raise PermissionError(f"Cannot write to CSV file '{filename}': {str(e)}. Check file permissions.")
        except OSError as e:
            raise OSError(f"File system error while exporting to '{filename}': {str(e)}")
        except Exception as e:
            raise RuntimeError(f"Unexpected error while exporting CSV: {str(e)}")

    def _filter_tasks_by_date(self, tasks: List[Task], target_date: datetime) -> List[Task]:
        target_date_only = target_date.date()
        return [
            task for task in tasks
            if (task.created_at.date() == target_date_only or
                (task.completed_at and task.completed_at.date() == target_date_only))
        ]

    def get_export_summary(self, tasks: List[Task]) -> Dict[str, Any]:
        if not isinstance(tasks, list):
            raise TypeError(f"Tasks must be a list, got {type(tasks)}")
        
        return {
            "total_tasks": len(tasks),
            "exportable": len(tasks) > 0,
            "estimated_size_kb": len(tasks) * 0.5,
            "fields_included": [
                'id', 'title', 'description', 'priority', 'status',
                'created_at', 'completed_at', 'project_id'
            ],
            "format": "CSV",
            "encoding": "UTF-8"
        }


class ExportService:
    """Service d'export vers différents formats (JSON, XML, Excel)"""
    
    SUPPORTED_FORMATS = ['json', 'xml', 'xlsx', 'excel']
    
    def __init__(self):
        self.export_history: List[Dict[str, Any]] = []
    
    def export_tasks(self, tasks: List[Task], filename: str, format_type: str = 'json', 
                     include_statistics: bool = True) -> bool:
        """
        Export des tâches vers différents formats
        
        Args:
            tasks: Liste des tâches à exporter
            filename: Nom du fichier de sortie
            format_type: Format d'export ('json', 'xml', 'xlsx', 'excel')
            include_statistics: Inclure les statistiques dans l'export
            
        Returns:
            bool: True si l'export a réussi
        """
        if not isinstance(tasks, list):
            raise TypeError(f"Tasks must be a list, got {type(tasks)}")
        
        if not isinstance(filename, str) or not filename.strip():
            raise ValueError("Filename cannot be empty")
        
        if not isinstance(format_type, str):
            raise TypeError(f"Format type must be a string, got {type(format_type)}")
        
        format_type = format_type.lower().strip()
        
        if format_type not in self.SUPPORTED_FORMATS:
            raise ValueError(f"Unsupported format: {format_type}. Supported formats: {self.SUPPORTED_FORMATS}")
        
        filename = filename.strip()
        
        # Détecter le format depuis l'extension si pas spécifié
        if format_type == 'excel' or filename.endswith('.xlsx'):
            format_type = 'xlsx'
        elif filename.endswith('.xml'):
            format_type = 'xml'
        elif filename.endswith('.json'):
            format_type = 'json'
        
        # Ajouter l'extension si manquante
        if not filename.endswith(f'.{format_type}'):
            filename += f'.{format_type}'
        
        try:
            if format_type == 'json':
                success = self._export_json(tasks, filename, include_statistics)
            elif format_type == 'xml':
                success = self._export_xml(tasks, filename, include_statistics)
            elif format_type == 'xlsx':
                success = self._export_excel(tasks, filename, include_statistics)
            else:
                raise ValueError(f"Unsupported format: {format_type}")
            
            # Enregistrer l'historique d'export
            self.export_history.append({
                "filename": filename,
                "format": format_type,
                "task_count": len(tasks),
                "include_statistics": include_statistics,
                "exported_at": datetime.now().isoformat(),
                "success": success
            })
            
            return success
            
        except Exception as e:
            self.export_history.append({
                "filename": filename,
                "format": format_type,
                "task_count": len(tasks),
                "include_statistics": include_statistics,
                "exported_at": datetime.now().isoformat(),
                "success": False,
                "error": str(e)
            })
            raise
    
    def _export_json(self, tasks: List[Task], filename: str, include_statistics: bool) -> bool:
        """Export vers JSON"""
        try:
            export_data = {
                "tasks": [task.to_dict() for task in tasks],
                "metadata": {
                    "total_tasks": len(tasks),
                    "export_format": "json",
                    "exported_at": datetime.now().isoformat(),
                    "include_statistics": include_statistics
                }
            }
            
            if include_statistics:
                export_data["statistics"] = self._generate_export_statistics(tasks)
            
            with open(filename, 'w', encoding='utf-8') as file:
                json.dump(export_data, file, indent=2, ensure_ascii=False)
            
            return True
            
        except Exception as e:
            raise RuntimeError(f"Error exporting to JSON: {str(e)}")
    
    def _export_xml(self, tasks: List[Task], filename: str, include_statistics: bool) -> bool:
        """Export vers XML"""
        try:
            # Créer l'élément racine
            root = ET.Element("TaskManagerExport")
            
            # Métadonnées
            metadata = ET.SubElement(root, "Metadata")
            ET.SubElement(metadata, "TotalTasks").text = str(len(tasks))
            ET.SubElement(metadata, "ExportFormat").text = "xml"
            ET.SubElement(metadata, "ExportedAt").text = datetime.now().isoformat()
            ET.SubElement(metadata, "IncludeStatistics").text = str(include_statistics)
            
            # Tâches
            tasks_element = ET.SubElement(root, "Tasks")
            for task in tasks:
                task_element = ET.SubElement(tasks_element, "Task")
                task_element.set("id", str(task.id))
                
                ET.SubElement(task_element, "Title").text = task.title
                ET.SubElement(task_element, "Description").text = task.description or ""
                ET.SubElement(task_element, "Priority").text = task.priority.value
                ET.SubElement(task_element, "Status").text = task.status.value
                ET.SubElement(task_element, "CreatedAt").text = task.created_at.isoformat()
                
                if task.completed_at:
                    ET.SubElement(task_element, "CompletedAt").text = task.completed_at.isoformat()
                
                if task.project_id:
                    ET.SubElement(task_element, "ProjectId").text = str(task.project_id)
            
            # Statistiques
            if include_statistics:
                stats = self._generate_export_statistics(tasks)
                stats_element = ET.SubElement(root, "Statistics")
                
                general_stats = ET.SubElement(stats_element, "GeneralStats")
                ET.SubElement(general_stats, "TotalTasks").text = str(stats["total_tasks"])
                ET.SubElement(general_stats, "CompletedTasks").text = str(stats["completed_tasks"])
                ET.SubElement(general_stats, "PendingTasks").text = str(stats["pending_tasks"])
                ET.SubElement(general_stats, "CompletionRate").text = str(stats["completion_rate"])
                
                # Répartition par priorité
                priority_stats = ET.SubElement(stats_element, "PriorityDistribution")
                for priority, count in stats["priority_distribution"].items():
                    priority_elem = ET.SubElement(priority_stats, "Priority")
                    priority_elem.set("type", priority)
                    priority_elem.text = str(count)
                
                # Répartition par statut
                status_stats = ET.SubElement(stats_element, "StatusDistribution")
                for status, count in stats["status_distribution"].items():
                    status_elem = ET.SubElement(status_stats, "Status")
                    status_elem.set("type", status)
                    status_elem.text = str(count)
            
            # Écrire le fichier XML
            tree = ET.ElementTree(root)
            tree.write(filename, encoding='utf-8', xml_declaration=True)
            
            return True
            
        except Exception as e:
            raise RuntimeError(f"Error exporting to XML: {str(e)}")
    
    def _export_excel(self, tasks: List[Task], filename: str, include_statistics: bool) -> bool:
        """Export vers Excel"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl library is required for Excel export. Install with: pip install openpyxl")
        
        try:
            # Créer le classeur
            wb = openpyxl.Workbook()
            
            # Feuille des tâches
            ws_tasks = wb.active
            ws_tasks.title = "Tasks"
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            # En-têtes
            headers = [
                "ID", "Title", "Description", "Priority", "Status",
                "Created At", "Completed At", "Project ID"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws_tasks.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            # Données des tâches
            for row, task in enumerate(tasks, 2):
                ws_tasks.cell(row=row, column=1, value=task.id)
                ws_tasks.cell(row=row, column=2, value=task.title)
                ws_tasks.cell(row=row, column=3, value=task.description or "")
                ws_tasks.cell(row=row, column=4, value=task.priority.value)
                ws_tasks.cell(row=row, column=5, value=task.status.value)
                ws_tasks.cell(row=row, column=6, value=task.created_at.strftime("%Y-%m-%d %H:%M:%S"))
                ws_tasks.cell(row=row, column=7, value=task.completed_at.strftime("%Y-%m-%d %H:%M:%S") if task.completed_at else "")
                ws_tasks.cell(row=row, column=8, value=task.project_id or "")
            
            # Ajuster la largeur des colonnes
            for column in ws_tasks.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_tasks.column_dimensions[column_letter].width = adjusted_width
            
            # Feuille des statistiques
            if include_statistics:
                ws_stats = wb.create_sheet(title="Statistics")
                stats = self._generate_export_statistics(tasks)
                
                # Titre
                ws_stats.cell(row=1, column=1, value="Task Statistics").font = Font(bold=True, size=16)
                ws_stats.cell(row=1, column=1).fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
                
                row = 3
                # Statistiques générales
                ws_stats.cell(row=row, column=1, value="General Statistics").font = Font(bold=True)
                row += 1
                ws_stats.cell(row=row, column=1, value="Total Tasks")
                ws_stats.cell(row=row, column=2, value=stats["total_tasks"])
                row += 1
                ws_stats.cell(row=row, column=1, value="Completed Tasks")
                ws_stats.cell(row=row, column=2, value=stats["completed_tasks"])
                row += 1
                ws_stats.cell(row=row, column=1, value="Pending Tasks")
                ws_stats.cell(row=row, column=2, value=stats["pending_tasks"])
                row += 1
                ws_stats.cell(row=row, column=1, value="Completion Rate")
                ws_stats.cell(row=row, column=2, value=f"{stats['completion_rate']}%")
                
                row += 3
                # Répartition par priorité
                ws_stats.cell(row=row, column=1, value="Priority Distribution").font = Font(bold=True)
                row += 1
                for priority, count in stats["priority_distribution"].items():
                    ws_stats.cell(row=row, column=1, value=priority.capitalize())
                    ws_stats.cell(row=row, column=2, value=count)
                    row += 1
                
                row += 2
                # Répartition par statut
                ws_stats.cell(row=row, column=1, value="Status Distribution").font = Font(bold=True)
                row += 1
                for status, count in stats["status_distribution"].items():
                    ws_stats.cell(row=row, column=1, value=status.replace("_", " ").title())
                    ws_stats.cell(row=row, column=2, value=count)
                    row += 1
                
                # Ajuster la largeur des colonnes
                for column in ws_stats.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    ws_stats.column_dimensions[column_letter].width = adjusted_width
            
            # Sauvegarder le fichier
            wb.save(filename)
            return True
            
        except Exception as e:
            raise RuntimeError(f"Error exporting to Excel: {str(e)}")
    
    def _generate_export_statistics(self, tasks: List[Task]) -> Dict[str, Any]:
        """Génère les statistiques pour l'export"""
        total_tasks = len(tasks)
        completed_tasks = len([t for t in tasks if t.status == Status.DONE])
        pending_tasks = len([t for t in tasks if t.status == Status.TODO])
        in_progress_tasks = len([t for t in tasks if t.status == Status.IN_PROGRESS])
        cancelled_tasks = len([t for t in tasks if t.status == Status.CANCELLED])
        
        completion_rate = (completed_tasks / total_tasks * 100) if total_tasks > 0 else 0
        
        priority_distribution = {
            "low": len([t for t in tasks if t.priority == Priority.LOW]),
            "medium": len([t for t in tasks if t.priority == Priority.MEDIUM]),
            "high": len([t for t in tasks if t.priority == Priority.HIGH]),
            "urgent": len([t for t in tasks if t.priority == Priority.URGENT])
        }
        
        status_distribution = {
            "todo": pending_tasks,
            "in_progress": in_progress_tasks,
            "done": completed_tasks,
            "cancelled": cancelled_tasks
        }
        
        return {
            "total_tasks": total_tasks,
            "completed_tasks": completed_tasks,
            "pending_tasks": pending_tasks,
            "in_progress_tasks": in_progress_tasks,
            "cancelled_tasks": cancelled_tasks,
            "completion_rate": round(completion_rate, 2),
            "priority_distribution": priority_distribution,
            "status_distribution": status_distribution,
            "generated_at": datetime.now().isoformat()
        }
    
    def get_export_history(self) -> List[Dict[str, Any]]:
        """Retourne l'historique des exports"""
        return self.export_history.copy()
    
    def clear_export_history(self) -> None:
        """Efface l'historique des exports"""
        self.export_history.clear()
    
    def get_supported_formats(self) -> List[str]:
        """Retourne la liste des formats supportés"""
        return self.SUPPORTED_FORMATS.copy()
    
    def is_format_supported(self, format_type: str) -> bool:
        """Vérifie si un format est supporté"""
        return format_type.lower() in self.SUPPORTED_FORMATS