import pytest
from unittest.mock import patch, Mock, mock_open
from datetime import datetime, timedelta
import csv
import io
import json
import os
import tempfile
from src.task_manager.services import EmailService, ReportService, ExportService
from src.task_manager.task import Task, Priority, Status


@pytest.mark.unit
class TestEmailService:
    """Tests du service email avec mocks"""

    def setup_method(self):
        self.email_service = EmailService("test.smtp.com", 587)

    def test_init_should_set_smtp_configuration(self):
        """Test initialisation définit configuration SMTP"""
        service = EmailService("custom.smtp.com", 465)
        
        assert service.smtp_server == "custom.smtp.com"
        assert service.port == 465
        assert service.sent_emails == []

    def test_send_task_reminder_with_valid_email_should_return_true(self):
        """Test envoi rappel avec email valide"""
        result = self.email_service.send_task_reminder(
            "user@example.com",
            "Tâche test",
            datetime.now()
        )
        
        assert result is True

    def test_send_task_reminder_should_record_sent_email(self):
        """Test envoi rappel enregistre email envoyé"""
        task_title = "Tâche importante"
        email = "user@example.com"
        due_date = datetime.now()
        
        self.email_service.send_task_reminder(email, task_title, due_date)
        
        assert len(self.email_service.sent_emails) == 1
        sent_email = self.email_service.sent_emails[0]
        assert sent_email["to"] == email
        assert task_title in sent_email["subject"]
        assert sent_email["type"] == "reminder"

    def test_send_task_reminder_with_empty_email_should_raise_error(self):
        """Test envoi rappel email vide lève erreur"""
        with pytest.raises(ValueError, match="Email cannot be empty"):
            self.email_service.send_task_reminder("", "Task", datetime.now())

    def test_send_task_reminder_with_invalid_email_format_should_raise_error(self):
        """Test envoi rappel email invalide lève erreur"""
        with pytest.raises(ValueError, match="Invalid email format: missing '@' symbol"):
            self.email_service.send_task_reminder("invalid-email", "Task", datetime.now())

    def test_send_task_reminder_with_missing_domain_should_raise_error(self):
        """Test envoi rappel domaine manquant lève erreur"""
        with pytest.raises(ValueError, match="Invalid email format: missing domain"):
            self.email_service.send_task_reminder("user@", "Task", datetime.now())

    def test_send_task_reminder_with_missing_extension_should_raise_error(self):
        """Test envoi rappel extension manquante lève erreur"""
        with pytest.raises(ValueError, match="Invalid email format: missing extension"):
            self.email_service.send_task_reminder("user@domain", "Task", datetime.now())

    def test_send_task_reminder_with_missing_local_part_should_raise_error(self):
        """Test envoi rappel partie locale manquante lève erreur"""
        with pytest.raises(ValueError, match="Invalid email format: missing local part"):
            self.email_service.send_task_reminder("@domain.com", "Task", datetime.now())

    def test_send_task_reminder_with_too_long_email_should_raise_error(self):
        """Test envoi rappel email trop long lève erreur"""
        long_email = "a" * 350 + "@domain.com"
        with pytest.raises(ValueError, match="Email too long.*Maximum: 320"):
            self.email_service.send_task_reminder(long_email, "Task", datetime.now())

    def test_send_task_reminder_with_invalid_characters_should_raise_error(self):
        """Test envoi rappel caractères invalides lève erreur"""
        with pytest.raises(ValueError, match="Invalid email format: contains invalid characters"):
            self.email_service.send_task_reminder("user space@domain.com", "Task", datetime.now())

    def test_send_task_reminder_with_empty_title_should_raise_error(self):
        """Test envoi rappel titre vide lève erreur"""
        with pytest.raises(ValueError, match="Task title cannot be empty"):
            self.email_service.send_task_reminder("user@domain.com", "", datetime.now())

    def test_send_task_reminder_with_invalid_date_type_should_raise_error(self):
        """Test envoi rappel date type invalide lève erreur"""
        with pytest.raises(TypeError, match="Due date must be a datetime object"):
            self.email_service.send_task_reminder("user@domain.com", "Task", "invalid")

    def test_send_completion_notification_with_valid_data_should_return_true(self):
        """Test envoi notification completion données valides"""
        result = self.email_service.send_completion_notification(
            "user@example.com",
            "Tâche terminée"
        )
        
        assert result is True

    def test_send_completion_notification_should_record_sent_email(self):
        """Test notification completion enregistre email"""
        email = "user@example.com"
        task_title = "Tâche terminée"
        
        self.email_service.send_completion_notification(email, task_title)
        
        assert len(self.email_service.sent_emails) == 1
        sent_email = self.email_service.sent_emails[0]
        assert sent_email["to"] == email
        assert task_title in sent_email["subject"]
        assert sent_email["type"] == "completion"

    def test_get_sent_emails_should_return_copy_of_emails(self):
        """Test récupération emails envoyés retourne copie"""
        self.email_service.send_task_reminder("user@domain.com", "Task", datetime.now())
        
        emails = self.email_service.get_sent_emails()
        emails.clear()
        
        assert len(self.email_service.sent_emails) == 1

    def test_clear_sent_emails_should_empty_emails_list(self):
        """Test nettoyage emails vide la liste"""
        self.email_service.send_task_reminder("user@domain.com", "Task", datetime.now())
        
        self.email_service.clear_sent_emails()
        
        assert len(self.email_service.sent_emails) == 0

    @pytest.mark.parametrize("email,should_pass", [
        ("user@example.com", True),
        ("test.user+tag@domain.org", True),
        ("a@b.co", True),
        ("invalid-email", False),
        ("user@", False),
        ("@domain.com", False),
        ("user@domain", False),
        ("user space@domain.com", False),
    ])
    def test_email_validation_various_formats(self, email, should_pass):
        """Test validation email divers formats"""
        if should_pass:
            result = self.email_service.send_task_reminder(email, "Task", datetime.now())
            assert result is True
        else:
            with pytest.raises(ValueError):
                self.email_service.send_task_reminder(email, "Task", datetime.now())


@pytest.mark.unit
class TestReportService:
    """Tests du service de rapports"""

    def setup_method(self):
        self.report_service = ReportService()
        self.sample_tasks = self._create_sample_tasks()

    def _create_sample_tasks(self):
        """Créer tâches échantillon pour tests"""
        tasks = []
        
        task1 = Task("Tâche TODO", "Description 1", Priority.HIGH)
        tasks.append(task1)
        
        task2 = Task("Tâche DONE", "Description 2", Priority.MEDIUM)
        task2.mark_completed()
        tasks.append(task2)
        
        task3 = Task("Tâche IN_PROGRESS", "Description 3", Priority.LOW)
        task3.status = Status.IN_PROGRESS
        tasks.append(task3)
        
        return tasks

    def test_generate_daily_report_with_tasks_should_return_report_structure(self):
        """Test génération rapport quotidien avec tâches"""
        report = self.report_service.generate_daily_report(self.sample_tasks)
        
        required_fields = [
            "report_date", "total_tasks", "tasks_for_date", "completed_today",
            "created_today", "completion_rate_today", "priority_breakdown",
            "status_breakdown", "generated_at", "summary"
        ]
        for field in required_fields:
            assert field in report

    def test_generate_daily_report_should_calculate_totals_correctly(self):
        """Test génération rapport calcul totaux"""
        report = self.report_service.generate_daily_report(self.sample_tasks)
        
        assert report["total_tasks"] == 3
        assert isinstance(report["completion_rate_today"], float)
        assert 0 <= report["completion_rate_today"] <= 100

    def test_generate_daily_report_should_include_priority_breakdown(self):
        """Test génération rapport inclut répartition priorités"""
        report = self.report_service.generate_daily_report(self.sample_tasks)
        
        priority_breakdown = report["priority_breakdown"]
        assert "urgent" in priority_breakdown
        assert "high" in priority_breakdown
        assert "medium" in priority_breakdown
        assert "low" in priority_breakdown

    def test_generate_daily_report_should_include_status_breakdown(self):
        """Test génération rapport inclut répartition statuts"""
        report = self.report_service.generate_daily_report(self.sample_tasks)
        
        status_breakdown = report["status_breakdown"]
        assert "todo" in status_breakdown
        assert "in_progress" in status_breakdown
        assert "done" in status_breakdown
        assert "cancelled" in status_breakdown

    def test_generate_daily_report_with_empty_tasks_should_raise_error(self):
        """Test génération rapport tâches vides lève erreur"""
        with pytest.raises(ValueError, match="Cannot generate report: no tasks found"):
            self.report_service.generate_daily_report([])

    def test_generate_daily_report_with_invalid_tasks_type_should_raise_error(self):
        """Test génération rapport type tâches invalide lève erreur"""
        with pytest.raises(TypeError, match="Tasks must be a list"):
            self.report_service.generate_daily_report("invalid")

    # @patch('src.task_manager.services.datetime')
    # def test_generate_daily_report_with_fixed_date_should_use_provided_date(self, mock_datetime):
    #     """Test génération rapport avec date fixe"""
    #     fixed_date = datetime(2024, 1, 15, 10, 0, 0)
    #     mock_datetime.now.return_value = fixed_date
        
    #     report = self.report_service.generate_daily_report(self.sample_tasks, fixed_date)
        
    #     assert report["report_date"] == fixed_date.isoformat()

    def test_generate_daily_report_with_invalid_date_type_should_raise_error(self):
        """Test génération rapport date type invalide lève erreur"""
        with pytest.raises(TypeError, match="Date must be a datetime object"):
            self.report_service.generate_daily_report(self.sample_tasks, "invalid")

    def test_export_tasks_csv_with_valid_tasks_should_return_true(self):
        """Test export CSV tâches valides"""
        with patch('builtins.open', mock_open()) as mock_file:
            result = self.report_service.export_tasks_csv(self.sample_tasks, "test.csv")
            
            assert result is True
            mock_file.assert_called_once()

    def test_export_tasks_csv_should_write_headers(self):
        """Test export CSV écrit en-têtes"""
        csv_content = io.StringIO()
        
        with patch('builtins.open', mock_open()) as mock_file:
            mock_file.return_value.write = csv_content.write
            self.report_service.export_tasks_csv(self.sample_tasks, "test.csv")
            
            mock_file.assert_called_once_with("test.csv", 'w', newline='', encoding='utf-8')

    def test_export_tasks_csv_with_empty_tasks_should_raise_error(self):
        """Test export CSV tâches vides lève erreur"""
        with pytest.raises(ValueError, match="Cannot export CSV: no tasks to export"):
            self.report_service.export_tasks_csv([], "test.csv")

    def test_export_tasks_csv_with_invalid_tasks_type_should_raise_error(self):
        """Test export CSV type tâches invalide lève erreur"""
        with pytest.raises(TypeError, match="Tasks must be a list"):
            self.report_service.export_tasks_csv("invalid", "test.csv")

    def test_export_tasks_csv_with_empty_filename_should_raise_error(self):
        """Test export CSV nom fichier vide lève erreur"""
        with pytest.raises(ValueError, match="Filename cannot be empty"):
            self.report_service.export_tasks_csv(self.sample_tasks, "")

    def test_export_tasks_csv_with_invalid_filename_type_should_raise_error(self):
        """Test export CSV nom fichier type invalide lève erreur"""
        with pytest.raises(TypeError, match="Filename must be a string"):
            self.report_service.export_tasks_csv(self.sample_tasks, None)

    def test_export_tasks_csv_should_add_csv_extension_if_missing(self):
        """Test export CSV ajoute extension si manquante"""
        with patch('builtins.open', mock_open()) as mock_file:
            self.report_service.export_tasks_csv(self.sample_tasks, "test")
            
            mock_file.assert_called_once_with("test.csv", 'w', newline='', encoding='utf-8')

    @patch('builtins.open', side_effect=PermissionError("Permission denied"))
    def test_export_tasks_csv_permission_denied_should_raise_error(self, mock_file):
        """Test export CSV permission refusée lève erreur"""
        with pytest.raises(PermissionError, match="Cannot write to CSV file.*Permission denied"):
            self.report_service.export_tasks_csv(self.sample_tasks, "readonly.csv")

    def test_get_export_summary_with_tasks_should_return_summary(self):
        """Test résumé export avec tâches"""
        summary = self.report_service.get_export_summary(self.sample_tasks)
        
        assert summary["total_tasks"] == 3
        assert summary["exportable"] is True
        assert summary["format"] == "CSV"
        assert summary["encoding"] == "UTF-8"
        assert "fields_included" in summary

    def test_get_export_summary_with_empty_tasks_should_indicate_not_exportable(self):
        """Test résumé export tâches vides indique non exportable"""
        summary = self.report_service.get_export_summary([])
        
        assert summary["total_tasks"] == 0
        assert summary["exportable"] is False

    def test_get_export_summary_with_invalid_type_should_raise_error(self):
        """Test résumé export type invalide lève erreur"""
        with pytest.raises(TypeError, match="Tasks must be a list"):
            self.report_service.get_export_summary("invalid")

    @pytest.mark.parametrize("task_count,expected_exportable", [
        (0, False),
        (1, True),
        (10, True),
        (100, True),
    ])
    def test_get_export_summary_exportable_status(self, task_count, expected_exportable):
        """Test statut exportable résumé"""
        tasks = [Task(f"Task {i}") for i in range(task_count)]
        
        summary = self.report_service.get_export_summary(tasks)
        
        assert summary["exportable"] == expected_exportable


@pytest.mark.unit
class TestExportService:
    """Tests du service d'export multi-format"""

    def setup_method(self):
        self.export_service = ExportService()
        self.sample_tasks = self._create_sample_tasks()
        self.temp_dir = tempfile.mkdtemp()

    def teardown_method(self):
        """Nettoyer les fichiers temporaires après chaque test"""
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def _create_sample_tasks(self):
        """Créer tâches échantillon pour tests"""
        tasks = []
        
        task1 = Task("Export Task 1", "Description export 1", Priority.HIGH)
        tasks.append(task1)
        
        task2 = Task("Export Task 2", "Description export 2", Priority.MEDIUM)
        task2.mark_completed()
        tasks.append(task2)
        
        task3 = Task("Export Task 3", "Description export 3", Priority.LOW)
        task3.status = Status.IN_PROGRESS
        tasks.append(task3)
        
        return tasks

    def _get_temp_file_path(self, filename):
        """Obtenir chemin fichier temporaire"""
        return os.path.join(self.temp_dir, filename)

    def test_init_should_initialize_export_service(self):
        """Test initialisation service d'export"""
        service = ExportService()
        
        assert service.export_history == []
        assert service.SUPPORTED_FORMATS == ['json', 'xml', 'xlsx', 'excel']

    def test_get_supported_formats_should_return_format_list(self):
        """Test récupération formats supportés"""
        formats = self.export_service.get_supported_formats()
        
        expected_formats = ['json', 'xml', 'xlsx', 'excel']
        assert formats == expected_formats

    def test_is_format_supported_should_validate_formats(self):
        """Test validation formats supportés"""
        assert self.export_service.is_format_supported('json') is True
        assert self.export_service.is_format_supported('xml') is True
        assert self.export_service.is_format_supported('xlsx') is True
        assert self.export_service.is_format_supported('excel') is True
        assert self.export_service.is_format_supported('pdf') is False
        assert self.export_service.is_format_supported('csv') is False

    def test_export_tasks_json_with_valid_data_should_succeed(self):
        """Test export JSON avec données valides"""
        filename = self._get_temp_file_path("test_export.json")
        
        result = self.export_service.export_tasks(
            self.sample_tasks, filename, "json", include_statistics=True
        )
        
        assert result is True
        assert os.path.exists(filename)

    def test_export_tasks_json_should_create_valid_json_structure(self):
        """Test export JSON crée structure JSON valide"""
        filename = self._get_temp_file_path("test_structure.json")
        
        self.export_service.export_tasks(
            self.sample_tasks, filename, "json", include_statistics=True
        )
        
        with open(filename, 'r', encoding='utf-8') as file:
            data = json.load(file)
        
        assert "tasks" in data
        assert "metadata" in data
        assert "statistics" in data
        assert len(data["tasks"]) == 3
        assert data["metadata"]["total_tasks"] == 3
        assert data["metadata"]["export_format"] == "json"

    def test_export_tasks_json_without_statistics_should_exclude_stats(self):
        """Test export JSON sans statistiques exclut stats"""
        filename = self._get_temp_file_path("test_no_stats.json")
        
        self.export_service.export_tasks(
            self.sample_tasks, filename, "json", include_statistics=False
        )
        
        with open(filename, 'r', encoding='utf-8') as file:
            data = json.load(file)
        
        assert "statistics" not in data
        assert data["metadata"]["include_statistics"] is False

    def test_export_tasks_xml_with_valid_data_should_succeed(self):
        """Test export XML avec données valides"""
        filename = self._get_temp_file_path("test_export.xml")
        
        result = self.export_service.export_tasks(
            self.sample_tasks, filename, "xml", include_statistics=True
        )
        
        assert result is True
        assert os.path.exists(filename)

    def test_export_tasks_xml_should_create_valid_xml_structure(self):
        """Test export XML crée structure XML valide"""
        filename = self._get_temp_file_path("test_structure.xml")
        
        self.export_service.export_tasks(
            self.sample_tasks, filename, "xml", include_statistics=True
        )
        
        # Vérifier que le fichier XML est lisible
        import xml.etree.ElementTree as ET
        tree = ET.parse(filename)
        root = tree.getroot()
        
        assert root.tag == "TaskManagerExport"
        assert root.find("Metadata") is not None
        assert root.find("Tasks") is not None
        assert root.find("Statistics") is not None
        
        tasks_element = root.find("Tasks")
        task_elements = tasks_element.findall("Task")
        assert len(task_elements) == 3

    def test_export_tasks_xml_without_statistics_should_exclude_stats(self):
        """Test export XML sans statistiques exclut stats"""
        filename = self._get_temp_file_path("test_xml_no_stats.xml")
        
        self.export_service.export_tasks(
            self.sample_tasks, filename, "xml", include_statistics=False
        )
        
        import xml.etree.ElementTree as ET
        tree = ET.parse(filename)
        root = tree.getroot()
        
        assert root.find("Statistics") is None
        metadata = root.find("Metadata")
        include_stats = metadata.find("IncludeStatistics")
        assert include_stats.text == "False"

    @pytest.mark.skipif(
        not hasattr(ExportService, '_export_excel') or 
        'openpyxl' not in str(ExportService._export_excel.__code__.co_names),
        reason="openpyxl not available"
    )
    def test_export_tasks_excel_with_valid_data_should_succeed(self):
        """Test export Excel avec données valides"""
        filename = self._get_temp_file_path("test_export.xlsx")
        
        result = self.export_service.export_tasks(
            self.sample_tasks, filename, "xlsx", include_statistics=True
        )
        
        assert result is True
        assert os.path.exists(filename)

    @pytest.mark.skipif(
        not hasattr(ExportService, '_export_excel') or 
        'openpyxl' not in str(ExportService._export_excel.__code__.co_names),
        reason="openpyxl not available"
    )
    def test_export_tasks_excel_should_create_workbook_with_sheets(self):
        """Test export Excel crée classeur avec feuilles"""
        filename = self._get_temp_file_path("test_workbook.xlsx")
        
        self.export_service.export_tasks(
            self.sample_tasks, filename, "xlsx", include_statistics=True
        )
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filename)
            
            assert "Tasks" in wb.sheetnames
            assert "Statistics" in wb.sheetnames
            
            tasks_sheet = wb["Tasks"]
            # Vérifier les en-têtes
            assert tasks_sheet.cell(row=1, column=1).value == "ID"
            assert tasks_sheet.cell(row=1, column=2).value == "Title"
            
            # Vérifier les données (3 tâches + 1 ligne d'en-tête = 4 lignes)
            assert tasks_sheet.max_row == 4
        except ImportError:
            pytest.skip("openpyxl not available for verification")

    def test_export_tasks_with_empty_tasks_list_should_succeed(self):
        """Test export avec liste tâches vide devrait réussir"""
        filename = self._get_temp_file_path("test_empty.json")
        
        result = self.export_service.export_tasks(
            [], filename, "json", include_statistics=True
        )
        
        assert result is True
        assert os.path.exists(filename)

    def test_export_tasks_with_invalid_tasks_type_should_raise_error(self):
        """Test export avec type tâches invalide lève erreur"""
        filename = self._get_temp_file_path("test_invalid.json")
        
        with pytest.raises(TypeError, match="Tasks must be a list"):
            self.export_service.export_tasks("invalid", filename, "json")

    def test_export_tasks_with_empty_filename_should_raise_error(self):
        """Test export avec nom fichier vide lève erreur"""
        with pytest.raises(ValueError, match="Filename cannot be empty"):
            self.export_service.export_tasks(self.sample_tasks, "", "json")

    def test_export_tasks_with_invalid_format_should_raise_error(self):
        """Test export avec format invalide lève erreur"""
        filename = self._get_temp_file_path("test_invalid_format.txt")
        
        with pytest.raises(ValueError, match="Unsupported format.*pdf"):
            self.export_service.export_tasks(self.sample_tasks, filename, "pdf")

    def test_export_tasks_should_add_extension_if_missing(self):
        """Test export ajoute extension si manquante"""
        filename_base = self._get_temp_file_path("test_no_ext")
        
        self.export_service.export_tasks(
            self.sample_tasks, filename_base, "json"
        )
        
        expected_filename = filename_base + ".json"
        assert os.path.exists(expected_filename)

    def test_export_tasks_should_detect_format_from_extension(self):
        """Test export détecte format depuis extension"""
        filename = self._get_temp_file_path("test_auto_format.xml")
        
        result = self.export_service.export_tasks(
            self.sample_tasks, filename, "json"  # Format spécifié JSON mais extension XML
        )
        
        assert result is True
        # Devrait créer un fichier XML basé sur l'extension
        import xml.etree.ElementTree as ET
        tree = ET.parse(filename)
        assert tree.getroot().tag == "TaskManagerExport"

    def test_export_tasks_should_record_export_history(self):
        """Test export enregistre historique"""
        filename = self._get_temp_file_path("test_history.json")
        
        self.export_service.export_tasks(
            self.sample_tasks, filename, "json", include_statistics=True
        )
        
        history = self.export_service.get_export_history()
        assert len(history) == 1
        
        export_record = history[0]
        assert export_record["filename"].endswith("test_history.json")
        assert export_record["format"] == "json"
        assert export_record["task_count"] == 3
        assert export_record["include_statistics"] is True
        assert export_record["success"] is True
        assert "exported_at" in export_record

    def test_export_tasks_should_record_failed_exports_in_history(self):
        """Test export enregistre échecs dans historique"""
        # Forcer une erreur en utilisant un répertoire inexistant
        invalid_filename = "/invalid/path/test_failure.json"
        
        with pytest.raises(Exception):
            self.export_service.export_tasks(
                self.sample_tasks, invalid_filename, "json"
            )
        
        history = self.export_service.get_export_history()
        assert len(history) == 1
        
        export_record = history[0]
        assert export_record["success"] is False
        assert "error" in export_record

    def test_get_export_history_should_return_copy(self):
        """Test récupération historique retourne copie"""
        filename = self._get_temp_file_path("test_copy.json")
        
        self.export_service.export_tasks(
            self.sample_tasks, filename, "json"
        )
        
        history = self.export_service.get_export_history()
        history.clear()
        
        # L'historique original ne devrait pas être affecté
        original_history = self.export_service.get_export_history()
        assert len(original_history) == 1

    def test_clear_export_history_should_empty_history(self):
        """Test nettoyage historique vide l'historique"""
        filename = self._get_temp_file_path("test_clear.json")
        
        self.export_service.export_tasks(
            self.sample_tasks, filename, "json"
        )
        
        self.export_service.clear_export_history()
        
        history = self.export_service.get_export_history()
        assert len(history) == 0

    @pytest.mark.parametrize("format_type,expected_format", [
        ("json", "json"),
        ("JSON", "json"),
        ("xml", "xml"),
        ("XML", "xml"),
        ("xlsx", "xlsx"),
        ("XLSX", "xlsx"),
        ("excel", "xlsx"),
        ("EXCEL", "xlsx"),
    ])
    def test_export_tasks_should_handle_case_insensitive_formats(self, format_type, expected_format):
        """Test export gère formats insensibles à la casse"""
        filename = self._get_temp_file_path(f"test_case_{format_type}.{expected_format}")
        
        result = self.export_service.export_tasks(
            self.sample_tasks, filename, format_type
        )
        
        assert result is True
        
        history = self.export_service.get_export_history()
        assert history[-1]["format"] == expected_format

    def test_generate_export_statistics_should_calculate_correctly(self):
        """Test génération statistiques export calcule correctement"""
        stats = self.export_service._generate_export_statistics(self.sample_tasks)
        
        assert stats["total_tasks"] == 3
        assert stats["completed_tasks"] == 1
        assert stats["pending_tasks"] == 1
        assert stats["in_progress_tasks"] == 1
        assert stats["cancelled_tasks"] == 0
        assert stats["completion_rate"] == 33.33
        
        assert stats["priority_distribution"]["high"] == 1
        assert stats["priority_distribution"]["medium"] == 1
        assert stats["priority_distribution"]["low"] == 1
        assert stats["priority_distribution"]["urgent"] == 0
        
        assert stats["status_distribution"]["todo"] == 1
        assert stats["status_distribution"]["in_progress"] == 1
        assert stats["status_distribution"]["done"] == 1
        assert stats["status_distribution"]["cancelled"] == 0

    def test_export_excel_without_openpyxl_should_raise_import_error(self):
        """Test export Excel sans openpyxl lève ImportError"""
        filename = self._get_temp_file_path("test_no_openpyxl.xlsx")
        
        # Simuler l'absence d'openpyxl
        with patch('src.task_manager.services.OPENPYXL_AVAILABLE', False):
            with pytest.raises(ImportError, match="openpyxl library is required"):
                self.export_service.export_tasks(
                    self.sample_tasks, filename, "xlsx"
                )

    def test_send_task_reminder_with_non_string_task_title_should_raise_error(self):
        """Test envoi rappel avec titre non-string lève erreur"""
        email_service = EmailService()
        with pytest.raises(TypeError, match="Task title must be a string"):
            email_service.send_task_reminder("user@domain.com", 123, datetime.now())

    def test_send_completion_notification_with_non_string_task_title_should_raise_error(self):
        """Test notification completion avec titre non-string lève erreur"""
        email_service = EmailService()
        with pytest.raises(TypeError, match="Task title must be a string"):
            email_service.send_completion_notification("user@domain.com", None)

    def test_send_task_reminder_with_non_string_email_should_raise_error(self):
        """Test envoi rappel avec email non-string lève erreur"""
        email_service = EmailService()
        with pytest.raises(TypeError, match="Email must be a string"):
            email_service.send_task_reminder(123, "Task title", datetime.now())

    def test_send_completion_notification_with_non_string_email_should_raise_error(self):
        """Test notification completion avec email non-string lève erreur"""
        email_service = EmailService()
        with pytest.raises(TypeError, match="Email must be a string"):
            email_service.send_completion_notification([], "Task title")


@pytest.mark.integration
class TestExportServiceRealFiles:
    """Tests d'intégration - création réelle de fichiers"""
    
    def setup_method(self):
        """Fixture : environnement de test"""
        self.temp_dir = tempfile.mkdtemp()
        self.export_service = ExportService()
        
        # Créer des tâches de test
        self.test_tasks = [
            Task("Tâche 1", "Description 1", Priority.HIGH),
            Task("Tâche 2", "Description 2", Priority.MEDIUM),
            Task("Tâche 3", "Description 3", Priority.LOW)
        ]
        self.test_tasks[0].mark_completed()  # Marquer la première comme terminée
    
    def teardown_method(self):
        """Nettoyage"""
        for file in os.listdir(self.temp_dir):
            os.unlink(os.path.join(self.temp_dir, file))
        os.rmdir(self.temp_dir)
    
    def test_real_json_export_creates_valid_file(self):
        """Test intégration : export JSON réel"""
        json_file = os.path.join(self.temp_dir, 'real_export.json')
        
        result = self.export_service.export_tasks(
            tasks=self.test_tasks,
            filename=json_file,
            format_type='json',
            include_statistics=True
        )
        
        assert result is True
        assert os.path.exists(json_file)
        
        # Vérifier le contenu
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        assert 'tasks' in data
        assert 'statistics' in data
        assert len(data['tasks']) == 3
        assert data['statistics']['total_tasks'] == 3
        assert data['statistics']['completed_tasks'] == 1
        assert data['statistics']['completion_rate'] == 33.33
    
    def test_real_xml_export_creates_valid_file(self):
        """Test intégration : export XML réel"""
        xml_file = os.path.join(self.temp_dir, 'real_export.xml')
        
        result = self.export_service.export_tasks(
            tasks=self.test_tasks,
            filename=xml_file,
            format_type='xml',
            include_statistics=True
        )
        
        assert result is True
        assert os.path.exists(xml_file)
        
        # Vérifier que le fichier n'est pas vide
        assert os.path.getsize(xml_file) > 0
        
        # Vérifier la structure XML basique
        with open(xml_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        assert '<TaskManagerExport>' in content
        assert '<Task id=' in content
        assert '<Statistics>' in content
        assert 'Tâche 1' in content
    
    def test_real_excel_export_creates_valid_file(self):
        """Test intégration : export Excel réel"""
        try:
            import openpyxl
            
            xlsx_file = os.path.join(self.temp_dir, 'real_export.xlsx')
            
            result = self.export_service.export_tasks(
                tasks=self.test_tasks,
                filename=xlsx_file,
                format_type='xlsx',
                include_statistics=True
            )
            
            assert result is True
            assert os.path.exists(xlsx_file)
            
            # Vérifier que le fichier n'est pas vide
            assert os.path.getsize(xlsx_file) > 0
            
            # Vérifier la structure Excel
            workbook = openpyxl.load_workbook(xlsx_file)
            assert 'Tasks' in workbook.sheetnames
            assert 'Statistics' in workbook.sheetnames
            
            # Vérifier le contenu de la feuille Tasks
            tasks_sheet = workbook['Tasks']
            assert tasks_sheet.cell(row=1, column=1).value == 'ID'
            assert tasks_sheet.cell(row=1, column=2).value == 'Title'
            assert tasks_sheet.cell(row=2, column=2).value == 'Tâche 1'
            
        except ImportError:
            pytest.skip("openpyxl not available")
    
    def test_export_history_tracking_with_real_files(self):
        """Test intégration : suivi de l'historique d'export"""
        # Exports vers plusieurs formats
        json_file = os.path.join(self.temp_dir, 'history_test.json')
        xml_file = os.path.join(self.temp_dir, 'history_test.xml')
        
        # Premier export
        self.export_service.export_tasks(self.test_tasks, json_file, 'json')
        
        # Deuxième export
        self.export_service.export_tasks(self.test_tasks, xml_file, 'xml')
        
        # Vérifier l'historique
        history = self.export_service.get_export_history()
        assert len(history) == 2
        
        # Vérifier les détails de l'historique
        json_export = next(h for h in history if h['format'] == 'json')
        xml_export = next(h for h in history if h['format'] == 'xml')
        
        assert json_export['filename'] == json_file
        assert xml_export['filename'] == xml_file
        assert json_export['success'] is True
        assert xml_export['success'] is True
        
        # Vérifier que les fichiers existent
        assert os.path.exists(json_file)
        assert os.path.exists(xml_file)